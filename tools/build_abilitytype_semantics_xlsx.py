# -*- coding: utf-8 -*-
"""Build exhaustive SpellAbilityType float1-3/int1-3 semantics workbook."""
from __future__ import annotations

import json
import os
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = r"D:\SteamLibrary\steamapps\common\Magicraft\_reverse_engineering"
OUT = os.path.join(BASE, "法术能力类型字段语义表.xlsx")
ENUM_PATH = os.path.join(BASE, "decompiled", "Assembly-CSharp", "SpellAbilityType.cs")
SPELL_JSON = os.path.join(BASE, "assets_export", "SpellConfig.json")
TEXT_JSON = os.path.join(BASE, "assets_export", "TextConfig_Spell.json")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
BODY_FONT = Font(name="Arial", size=10)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT = PatternFill("solid", fgColor="F2F2F2")
CONF_FILL = {
    "high": PatternFill("solid", fgColor="C6EFCE"),
    "medium": PatternFill("solid", fgColor="FFEB9C"),
    "low": PatternFill("solid", fgColor="FCE4D6"),
    "unknown": PatternFill("solid", fgColor="FFC7CE"),
}

# Semantics mined from decompiled code + TextConfig + SpellConfig samples.
# Format: id -> (category, float1, float2, float3, int1, int2, int3, evidence, confidence, notes)
# Unused slots: "-" ; Unknown: "unknown"
SEMANTICS: dict[int, tuple] = {
    0: ("none", "-", "-", "-", "-", "-", "-", "SpellAbilityType.cs; SpellConfig.json", "high", "占位/无效类型"),
    1000: ("missile", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "high", "空想之术空槽"),
    1001: ("missile", "(Lv2+)命中后易伤伤害加成%", "(Lv2+)易伤持续(秒)", "-", "(Lv2+)启用相关标志", "(Lv2+)未知(配置10/30)", "-", "Spell1001Bullet.cs; UnitProperty*.cs; TextConfig 7110012", "high", "f1/f2高置信; i1/i2偏低"),
    1002: ("missile", "可吸收/滚球HP池", "配置有值但代码未读(50-60)", "-", "-", "-", "-", "Spell1002RollBall.cs; SpellTools.AttackRollball", "high", "f2疑似残留"),
    1003: ("missile", "速度衰减插值系数", "开始追踪前的最低速度比例", "-", "-", "-", "-", "Spell1003ButterFlySystem.cs", "high", ""),
    1004: ("missile", "unknown(配置4)", "unknown(配置2)", "-", "(Lv2+)穿透次数(+折射转换)", "-", "-", "Spell1004Laser.cs; TextConfig 7110042", "high", "i1高置信; f1/f2未知"),
    1005: ("missile", "结束时右侧法术蓝耗比例%", "unknown(配置2)", "-", "-", "-", "-", "SpellShootGroup.cs; SpellSpawnParamsProcessor.cs", "high", ""),
    1006: ("missile", "(Lv2+)自我复制间隔(秒)", "随机散射角±度", "-", "(Lv2+)启用标志", "-", "-", "Spell1006GhostFireSystem.cs", "high", ""),
    1007: ("missile", "半径增长速度(m/s)", "配置有值代码未读", "配置有值代码未读", "-", "-", "-", "Spell1007BlackHole.cs", "high", "f2/f3疑似未用"),
    1008: ("missile", "(Lv2+)击杀生成体积比例", "-", "-", "(Lv2+)启用击杀爆炸", "-", "-", "Spell1008ArcaneExplosion.cs", "high", ""),
    1009: ("missile", "-", "-", "-", "配置有值(3)但描述/代码未用", "-", "-", "Spell1009BackMP.cs", "medium", "回蓝=本次施法全额消耗"),
    1010: ("missile", "蛇形轨迹最大长度(m)", "-", "-", "-", "-", "-", "Spell1010SnakeWalkSystem.cs", "high", ""),
    1011: ("missile", "unknown(1.5)", "unknown(0.1)", "unknown(10)", "unknown(12/24/48)", "-", "-", "SpellConfig.json; TextConfig无占位符", "unknown", "需对照Authoring深入"),
    1012: ("missile", "配置残留(0.33)", "配置残留(2)", "配置残留(2)", "-", "-", "-", "SpellSpawnParamsProcessor.cs(行为硬编码)", "medium", "行为多硬编码"),
    1013: ("missile", "-", "-", "-", "(Lv2+)额外小陨石数", "-", "运行时陨石索引(Int3)", "Spell1013Meteor*.cs", "high", ""),
    1014: ("missile", "减速至停止的插值系数", "-", "-", "-", "-", "-", "Spell1014Rainbow*.cs", "high", ""),
    1015: ("missile", "嵌套施法蓝耗比例%", "unknown(0.5)", "-", "嵌套施法次数", "嵌套施法伤害%", "-", "Spell1015ArcaneNova.cs; SpellShootGroup.cs", "high", ""),
    1016: ("missile", "结束爆炸半径倍率×radius", "-", "-", "(Lv2+)启用结束爆炸", "-", "-", "Spell1016Dash*.cs", "high", ""),
    1017: ("missile", "-", "-", "-", "每级升级所需击杀数", "-", "-", "Spell1017DeathAdder*.cs; SlotData.cs", "high", ""),
    1018: ("missile", "(Lv2+)单体伤害加成%", "链式重定向延迟×DPS间隔", "-", "最大传导/链数", "-", "-", "Spell1018ThunderAura*.cs", "high", ""),
    1019: ("missile", "unknown(0.033)", "-", "-", "unknown(8)", "-", "运行时:0=持续附着,非0=普通移动", "SpellMoveJob.cs; Spell1019*", "medium", "i3为运行时状态"),
    1020: ("missile", "消耗玩家金币比例%", "每金币伤害加成因子", "-", "unknown(8)", "-", "运行时金币状态(Int3)", "Spell1020ManaCoin*.cs; CanShootSpellUtils.cs", "high", ""),
    1021: ("missile", "unknown(90)", "unknown(1.8/10)", "运行时下落水平速度", "unknown(8)", "-", "运行时翻转Y(±1)", "Spell1021*.cs", "medium", "部分为运行时写入"),
    1022: ("missile", "-", "unknown(30)", "-", "内置穿透次数", "-", "-", "Spell1022Boomerang.cs", "high", ""),
    1023: ("missile", "索敌/攻击距离(m)", "unknown(30)", "-", "unknown(16)", "-", "-", "Spell1023JudgementBlade*.cs", "high", "f1高置信"),
    1024: ("missile", "unknown(1)", "效果半径+%每秒", "伤害+%每秒", "unknown(16)", "-", "-", "Spell1024GiantBubble*.cs", "high", "f2/f3高置信"),
    1025: ("missile", "基础攻击距离(m)", "每秒蓝耗(与持续相关)", "施法中伤害爬升%/秒", "unknown(10)", "-", "-", "Spell1025DragonBreath*.cs; SpellConfig.GetDes", "high", ""),
    1026: ("missile", "蓄力中暴击率+%每秒", "暴击≥float3时最终伤害×%", "触发加成的暴击阈值%", "unknown(10)", "-", "-", "Spell1026*.cs", "high", ""),
    1027: ("missile", "蓄力中基础伤害+%每秒", "最大蓄力时间(秒)", "unknown(100)", "unknown(10)", "-", "-", "Spell1027SuperNova.cs", "high", "f1/f2高置信"),
    1028: ("missile", "模板残留?", "模板残留?", "模板残留?", "普通箭次数后放大箭", "大箭数量", "运行时标志", "Spell1028MrBingArrow.cs; SlotData.cs", "high", "i1/i2高置信"),
    1029: ("missile", "每单位MP换算系数", "每float1 MP的额外基础伤害", "每float1 MP的额外持续(秒)", "模板残留?", "模板残留?", "-", "Spell1029DimensionTraveller*.cs", "high", ""),
    1030: ("missile", "附着时MP汲取/秒(描述)", "附着DPS(描述)", "附着伤害系数(伤害×间隔×f3)", "模板残留", "模板残留", "-", "Spell1030Harpoons.cs", "medium", "f3代码确认"),
    1031: ("missile", "模板残留?", "模板残留?", "模板残留?", "命中N次后+1鱼叉", "-", "-", "TextConfig 7110311", "high", "i1高置信"),
    2001: ("summon", "-", "-", "-", "配置残留ID?", "-", "-", "SpellSpawnParamsProcessor.cs", "medium", "召唤物属性看UnitConfig"),
    2002: ("summon", "(Lv2+)DPS=MaxMP×float1%", "腿部回血量", "-", "腿数量", "(Lv2+)标志", "-", "Spell2002*.cs; CanShootSpellUtils.cs", "high", ""),
    2003: ("summon", "-", "-", "-", "配置残留?", "-", "-", "SpellConfig.json", "medium", ""),
    2004: ("summon", "-", "unknown(0.5)", "接触伤害=MaxHP×float3%", "unknown(3-5)", "-", "-", "Spell2004TriggerJob.cs", "high", "f3高置信"),
    2005: ("summon", "回蓝=法杖回蓝×float1%", "-", "-", "-", "-", "-", "CanShootSpellUtils.cs", "high", ""),
    2006: ("summon", "每击杀暴击/MaxHP+%", "-", "灵魂炸弹伤害=MaxHP×float3", "运行时击杀计数", "-", "-", "Spell2006System.cs; SlotData.cs", "high", ""),
    2007: ("summon", "unknown(22)", "自伤间隔(秒)", "unknown(0.5)", "受伤时刷虫数", "死亡时刷虫数", "-", "Spell2007SuicideBugNest*.cs", "high", ""),
    2008: ("summon", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", "枚举有定义但配置未导出"),
    2009: ("summon", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", "枚举有定义但配置未导出"),
    3001: ("enhance", "蓝耗×(float1%)^(额外齐射-1)", "(Lv2+)额外/未知", "-", "+同时施法数", "每额外齐射散射+°", "-", "SpellShootData.cs; Wand.cs", "high", ""),
    3002: ("enhance", "多重射击间距", "-", "-", "额外施法次数", "-", "-", "SpellShootData.GetSpellMultiShootData", "high", ""),
    3003: ("enhance", "模板残留", "模板残留", "模板残留", "+同时施法数", "-", "-", "SpellShootData.cs; Wand.cs", "high", "i1高置信"),
    3004: ("enhance", "粘液持续(秒)", "移速×%", "法术飞行速度×%", "-", "-", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3005: ("enhance", "毒液持续(秒)", "-", "-", "unknown(配置80-380)", "施加毒层数", "-", "SpellBase.cs; ShootSpellUtils.cs", "high", "f1/i2高置信"),
    3006: ("enhance", "-", "-", "-", "+穿透次数", "-", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3007: ("enhance", "链间距/宽度?(0.2)", "-", "-", "链条路径伤害", "-", "-", "CanShootSpellUtils.cs; SpellConfig.GetDes", "high", "i1高置信"),
    3008: ("enhance", "悬浮持续(秒)", "-", "-", "-", "-", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3009: ("enhance", "+环绕持续(秒)", "+环绕移动速度", "环绕半径(m)", "-", "-", "-", "SpellShootData.GetSpellRotationInfo", "high", ""),
    3010: ("enhance", "+飞行速度", "+持续(秒)", "鼠标追踪插值速度", "-", "-", "-", "SpellShootData.cs; SpellBase.cs", "high", ""),
    3011: ("enhance", "追踪角速度(°/m)", "-", "-", "-", "-", "-", "SpellShootData.cs; SpellBase.cs", "high", ""),
    3012: ("enhance", "每次反弹+持续(秒)", "-", "-", "+反弹次数", "-", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3013: ("enhance", "模板残留", "模板残留", "模板残留", "分裂数量", "分裂伤害%(描述;ECS常硬编码0.33)", "-", "SpellBase.cs; ShootSpellUtils.cs", "high", ""),
    3014: ("enhance", "冰冻持续(秒)", "-", "-", "-", "-", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3015: ("enhance", "召唤物HP损失/秒", "召唤物攻速+%", "召唤物移速+%", "残留ID?", "死亡寄生数", "unknown(4)", "SpellBase.cs; CanShootSpellUtils.cs", "high", ""),
    3101: ("enhance", "结束落雷半径(m)", "结束落雷伤害%", "触发概率%", "unknown", "unknown", "-", "SpellBase.cs; ShootSpellUtils.cs", "high", "f1-f3高置信"),
    3102: ("enhance", "伤害+%(加算AddRatio)", "-", "-", "-", "-", "-", "CanShootSpellUtils.GetDamage", "high", ""),
    3103: ("enhance", "持续时间+(秒)", "-", "-", "-", "-", "-", "SpellShootData.cs", "high", ""),
    3104: ("enhance", "unknown/残留", "持续时间×%", "伤害×%(亦影响召唤HP)", "unknown/残留", "-", "-", "SpellShootData.cs; CanShootSpellUtils.cs", "high", "f2/f3高置信"),
    3105: ("enhance", "-", "-", "-", "-", "-", "-", "SpellShootData.GetSpellCriticalChance", "high", "用criticalChance字段,不用float/int槽"),
    3106: ("enhance", "飞行速度+", "-", "-", "-", "-", "-", "SpellShootData.cs; SpellBase.cs", "high", ""),
    3107: ("enhance", "效果半径+%", "-", "-", "-", "-", "-", "SpellShootData.cs", "high", ""),
    3108: ("enhance", "召唤物HP回复/秒", "召唤物MaxHP+%", "-", "-", "-", "-", "SpellShootData.cs; SpellBase.cs", "high", ""),
    3109: ("enhance", "配置有值疑似未用", "配置有值疑似未用", "-", "-", "-", "-", "Wand.cs(模拟逻辑)", "medium", "行为靠mimicSpellID"),
    3110: ("enhance", "路径每跳伤害", "跳伤间隔(秒)", "-", "unknown(2)", "-", "-", "CanShootSpellUtils.GetLifeLineDamage", "high", ""),
    3111: ("enhance", "伤害+%(加算)", "燃烧持续(秒)", "-", "燃烧MaxHP%/秒", "-", "-", "CanShootSpellUtils.cs; ShootSpellUtils.cs", "high", ""),
    3112: ("enhance", "吸引/拖拽半径(m)", "吸引力", "-", "暴击拖拽伤害%(命中)", "吸引目标数", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3113: ("enhance", "最终半径×%", "每损失1m半径最终伤害+%", "-", "unknown", "unknown", "-", "SpellShootData.cs", "high", ""),
    3114: ("enhance", "跟随角速度(°/m)", "unknown", "-", "unknown", "unknown", "-", "SpellShootData.cs; SpellBase.cs", "high", "f1高置信"),
    3115: ("enhance", "融合召唤属性+%", "unknown", "-", "最大融合等级/次数", "unknown", "-", "CanShootSpellUtils.cs; Spell2001Summon1.cs", "high", ""),
    3116: ("enhance", "标志型(配置值未读)", "标志型", "-", "标志型", "标志型", "-", "SpellShootData.GetSpellIsEndTeleport", "high", "有无即可触发结束传送"),
    3117: ("enhance", "随机半径下限", "随机半径上限", "+持续(秒)/轨道半径种子", "unknown", "unknown", "-", "SpellBase.cs; SpellShootData.cs", "high", ""),
    3118: ("enhance", "召唤物MaxHP×%", "死亡爆炸半径(m)", "死亡爆炸HP伤害%", "即死HP阈值%", "unknown", "-", "CanShootSpellUtils.cs", "high", ""),
    3119: ("enhance", "下落冲击/爆炸半径(m)", "unknown(1)", "-", "-", "-", "-", "SpellShootData.GetSpellFallExplosionRadius", "high", ""),
    3120: ("enhance", "-", "-", "-", "灵魂精华/进阶技能等级", "-", "-", "SpellShootData.cs; CanShootSpellUtils.cs", "high", ""),
    3121: ("enhance", "折射搜索半径", "伤害×%(乘算MulRatio)", "-", "折射次数", "-", "-", "SpellShootData.GetSpellRefractionInfo; CanShootSpellUtils", "high", ""),
    3122: ("enhance", "死后继续战斗时长(秒)", "-", "-", "-", "-", "-", "CanShootSpellUtils.cs", "high", ""),
    3123: ("enhance", "伤害+%(加算)", "-", "-", "-", "-", "-", "CanShootSpellUtils.cs", "high", "蓝耗另走mpCostMulDiv"),
    3124: ("enhance", "体积/尺寸+%", "-", "-", "-", "-", "-", "SpellShootData.GetSpellVolumeRatio", "high", ""),
    3125: ("enhance", "反向施法额外速度", "-", "-", "-", "-", "-", "CanShootSpellUtils.GetReverseShootState", "high", ""),
    3126: ("enhance", "标志型(配置100未读)", "-", "-", "-", "-", "-", "Wand.cs(+1等级逻辑)", "high", ""),
    3127: ("enhance", "跟随进下一房间概率%", "获得召唤者MaxHP%", "-", "-", "-", "-", "CanShootSpellUtils.cs", "high", ""),
    3128: ("enhance", "速度×%", "每损失1速度+持续(秒)", "-", "-", "-", "-", "SpellShootData.cs", "high", ""),
    3129: ("enhance", "虚空爆炸半径(m)", "MaxHP伤害分摊%", "即杀HP阈值%", "-", "-", "-", "SpellBase.cs; ShootSpellUtils.cs", "high", ""),
    3130: ("enhance", "传送半径(m)/半衰期传送", "传送后+持续(秒)", "传送后+速度", "-", "-", "-", "SpellShootData.GetHalfLifeTeleportData", "high", ""),
    3201: ("trigger", "unknown(1)", "-", "-", "unknown(100)", "继承父法术伤害%", "-", "SpellTriggerController.cs; ShootSpellUtils.cs", "high", "i2高置信"),
    3202: ("trigger", "unknown", "-", "-", "unknown(100)", "分裂射击数", "伤害×%", "SpellTriggerController.cs", "high", "i2/i3高置信"),
    3203: ("trigger", "触发距离比例(蓝耗/距离)", "-", "-", "子组蓝耗%", "unknown(3)", "-", "SpellTriggerController.cs", "high", ""),
    3204: ("trigger", "缠绕/环绕半径", "缠绕法术额外速度", "-", "unknown(100)", "缠绕生成数量", "-", "SpellTriggerController.cs", "high", ""),
    3205: ("trigger", "unknown", "-", "-", "蓝耗%", "射速:间隔=1/int2秒", "-", "SpellTriggerController.cs; SpellConfig.GetDes", "high", "i1/i2高置信"),
    4001: ("passive", "-", "-", "-", "-", "-", "-", "SpellConfig.json全0", "medium", "献祭球"),
    4002: ("passive", "最大MP+%", "-", "-", "-", "-", "-", "Wand.cs; WandConfig.GetExtraMaxMP", "high", ""),
    4003: ("passive", "MP回复+/秒", "-", "-", "-", "-", "-", "Wand.cs; WandConfig.GetExtraMPRecovery", "high", ""),
    4004: ("passive", "-", "蓄力间隔(秒)", "-", "最大蓄力层数", "-", "-", "Wand.cs", "high", ""),
    4005: ("passive", "自动法杖MP回复×%", "-", "-", "-", "-", "-", "Wand.cs", "high", ""),
    4006: ("passive", "-", "-", "-", "-", "-", "-", "Wand.cs", "high", "用coolDownRatio字段"),
    4007: ("passive", "-", "最终伤害×%", "-", "-", "-", "-", "Wand.cs", "high", ""),
    4008: ("passive", "回声施法概率%(加算)", "免蓝概率%", "unknown", "unknown", "-", "-", "Wand.cs", "high", "f1/f2高置信"),
    4009: ("passive", "最大MP×%", "MP回复-/秒", "-", "-", "-", "-", "Wand.cs; TextConfig 7140091", "high", "SpellConfig导出缺失此行"),
    4010: ("passive", "标志/等级型", "标志型", "标志型", "标志型", "-", "-", "Wand.cs", "high", "全域共享强化"),
    4011: ("passive", "标志型", "标志型", "标志型", "标志型", "标志型", "-", "Wand.cs(passiveEqualAngleDistribution)", "high", "等分角度"),
    4012: ("passive", "unknown", "挡1伤耗MP", "反击半径(m)", "反击伤害=耗蓝×int1", "最大MP+", "-", "Wand.cs; SpellInfoFormatters.cs", "high", "保护伞"),
    4013: ("passive", "unknown", "持有者移速+%", "unknown", "锤子伤害(formatter)", "unknown", "-", "Wand.cs; SpellInfoFormatters.cs", "high", "符文之锤"),
    4014: ("passive", "命中时基础蓝耗/秒", "持续命中速率爬升%/秒", "unknown", "基础DPS/随速率加算伤害", "unknown", "-", "Spell4014*.cs; SpellInfoFormatters.cs", "high", "激光束水晶"),
    4015: ("passive", "每移动1m充能", "模板残留", "模板残留", "模板残留", "模板残留", "-", "Wand.cs", "high", "蓄能槽延伸-移动"),
    4016: ("passive", "每站立1秒充能", "模板残留", "模板残留", "模板残留", "模板残留", "-", "Wand.cs", "high", "蓄能槽延伸-站立"),
    4017: ("passive", "每秒充能", "模板残留", "模板残留", "模板残留", "模板残留", "-", "Wand.cs", "high", "蓄能槽延伸-时间"),
    4018: ("passive", "每次法杖施法充能", "模板残留", "模板残留", "模板残留", "模板残留", "-", "Wand.cs", "high", "蓄能槽延伸-施法"),
    4019: ("passive", "每+1飞剑上限耗MaxMP", "飞剑伤害(formatter)", "unknown", "unknown", "unknown", "-", "SpellInfoFormatters.cs; Wand.cs", "high", "彼岸致命之刃"),
    4020: ("passive", "材料标志型", "材料标志型", "材料标志型", "材料标志型", "材料标志型", "-", "Wand.cs; UICompound.cs", "high", "法术胚胎"),
    4021: ("passive", "每占用格MP回复+%", "每占用格MaxMP+%", "unknown", "unknown", "unknown", "-", "Wand.cs; SlotData.cs", "high", "魔力卷须;格数=specialInt+1"),
    4022: ("passive", "随机生成半径", "每发角度偏移", "unknown", "unknown", "unknown", "-", "Wand.cs", "high", ""),
    4023: ("passive", "unknown", "unknown", "每1点MP回复→蓄能槽充能+%", "unknown", "unknown", "-", "Wand.cs", "high", "f3高置信"),
    4024: ("passive", "unknown(48)", "-", "unknown(2)", "unknown", "unknown", "-", "Wand.cs; SpellSpawnParamsProcessor.cs", "low", "戴夫鱼叉启用"),
    4025: ("passive", "每层红符文伤害", "AOE斩击阈值/标志(描述25%)", "超级斩击加成系数", "运行时符文计数", "运行时符文等级", "-", "Spell4025System.cs; SpellInfoFormatters.cs", "high", "红符文"),
    4026: ("passive", "每层绿符文伤害", "冲击获取间隔(秒)", "unknown", "最大存储层数(描述5)", "生成/扇形数量", "运行时直接爆炸标志", "Spell4026*.cs; SpellInfoFormatters.cs", "high", "绿符文"),
    4027: ("passive", "每层蓝符文伤害", "每次触发耗MP", "运行时MP回补量(Float3)", "每次触发导弹数", "运行时受伤增幅(×1/10000)", "运行时已充能标志", "Spell4027*.cs; SpellInfoFormatters.cs", "high", "蓝符文"),
    9001: ("monster", "弧线参数/类upSpeed", "类gravity", "-", "-", "-", "-", "SpellConfig.json; 亦用upSpeed/gravity字段", "medium", "抛物线弹道组件"),
    9002: ("monster", "爆炸伤害缩放max(f1,1)", "爆炸基础伤害覆盖", "-", "-", "-", "-", "Spell9002BoBoBombSystem.cs", "high", ""),
    9003: ("monster", "-", "-", "-", "-", "-", "-", "Spell9003LongTrailSystem.cs", "medium", "表现组件"),
    9004: ("monster", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "medium", ""),
    9005: ("monster", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "medium", ""),
    9006: ("monster", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "medium", ""),
    9007: ("monster", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", ""),
    9008: ("monster", "正弦振幅(°)(常运行时写)", "正弦频率", "正弦振幅缩放", "-", "-", "-", "Spell9008SingWaveSpeedJob.cs", "high", "JSON常为0,生成时写入"),
    9009: ("monster", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", ""),
    9010: ("monster", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "medium", ""),
    9011: ("monster", "目标X(运行时)", "目标Y(运行时)", "旋转速度(运行时)", "-", "-", "-", "Spell9011RotateArrowJob.cs", "high", "JSON常为0"),
    9012: ("monster", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", ""),
    9013: ("monster", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", ""),
    9014: ("monster", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "medium", ""),
    9015: ("monster", "-", "-", "-", "-", "-", "-", "SpellConfig.json", "medium", ""),
    9016: ("monster", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "无SpellConfig行", "SpellAbilityType.cs", "unknown", ""),
}

# Monster bullets 9050-9065 fill defaults
for mid in range(9050, 9066):
    if mid not in SEMANTICS:
        has_row_note = "无SpellConfig行" if mid in (9050, 9051, 9057, 9059) else "-"
        conf = "unknown" if has_row_note.startswith("无") else "medium"
        SEMANTICS[mid] = (
            "monster",
            has_row_note if has_row_note.startswith("无") else "-",
            has_row_note if has_row_note.startswith("无") else "-",
            has_row_note if has_row_note.startswith("无") else "-",
            has_row_note if has_row_note.startswith("无") else "-",
            has_row_note if has_row_note.startswith("无") else "-",
            has_row_note if has_row_note.startswith("无") else "-",
            "SpellAbilityType.cs; SpellConfig.json",
            conf,
            "怪物弹幕组件;数值多走damage/speed/duration等主字段或生成时写入",
        )

CATEGORY_CN = {
    "none": "无效/占位",
    "missile": "主动弹幕(1000段)",
    "summon": "召唤(2000段)",
    "enhance": "强化/修饰(3000/3100段)",
    "trigger": "触发器(3200段)",
    "passive": "法杖核心/被动(4000段)",
    "monster": "怪物弹道组件(9000段)",
}


def load_enum() -> dict[int, str]:
    text = open(ENUM_PATH, encoding="utf-8").read()
    return {int(v): k for k, v in re.findall(r"(\w+)\s*=\s*(\d+)", text)}


def sample_values(spells: list[dict]) -> dict[int, str]:
    by = defaultdict(list)
    for s in spells:
        by[s["abilityType"]].append(s)
    out = {}
    for ab, rows in by.items():
        # prefer level 1, else first
        rows = sorted(rows, key=lambda r: r.get("level", 1))
        parts = []
        for r in rows[:3]:
            bits = [f"id={r['id']}(Lv{r.get('level',1)})"]
            for f in ("float1", "float2", "float3", "int1", "int2", "int3"):
                v = r.get(f)
                if v:
                    bits.append(f"{f}={v}")
            parts.append("; ".join(bits))
        out[ab] = " | ".join(parts)
    return out


def spell_names(spells: list[dict], texts: list[dict]) -> dict[int, str]:
    """Representative Chinese name per abilityType (from lowest-level spell)."""
    tmap = {t["id"]: t for t in texts}
    by = defaultdict(list)
    for s in spells:
        by[s["abilityType"]].append(s)
    out = {}
    for ab, rows in by.items():
        r = sorted(rows, key=lambda x: x.get("level", 1))[0]
        t = tmap.get(r["id"] + 7_000_000, {})
        out[ab] = t.get("chineseS", "") or ""
    return out


def style_header(ws, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"


def autosize(ws, max_w=56):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:80]:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(8, min(length + 2, max_w))


def main():
    enum_map = load_enum()
    spells = json.load(open(SPELL_JSON, encoding="utf-8"))
    texts = json.load(open(TEXT_JSON, encoding="utf-8"))
    samples = sample_values(spells)
    names = spell_names(spells, texts)
    in_config = {s["abilityType"] for s in spells}

    wb = Workbook()

    # --- README ---
    ws = wb.active
    ws.title = "README"
    readme = [
        ("《魔法工艺》SpellAbilityType 通用字段语义穷举表",),
        ("",),
        ("覆盖范围", "SpellAbilityType 枚举全部 151 项（含 SpellConfig 中出现的 139 种 + 枚举有但配置缺失的 12 种）"),
        ("通用槽位", "float1 / float2 / float3 / int1 / int2 / int3 —— 语义完全由 abilityType 决定"),
        ("证据来源", "反编译源码(SpellShootData/CanShootSpellUtils/Wand/SpellTriggerController/各Spell*System) + TextConfig_Spell + SpellConfig.json"),
        ("置信度", "high=代码/描述直接确认; medium=间接或未使用但可判定; low=部分猜测; unknown=证据不足"),
        ("符号", '"-" = 该槽位未使用; "模板残留" = JSON有值但运行时代码未读; "运行时" = 生成后由系统写入而非配置表驱动'),
        ("",),
        ("工作表说明", ""),
        ("AbilityType语义", "主表：每种 abilityType 的六槽位含义、分类、置信度、证据、配置样本"),
        ("按等级样本值", "SpellConfig 中每条记录的 float/int 原始值（便于对照）"),
        ("置信度统计", "汇总"),
        ("",),
        ("生成脚本", "tools/build_abilitytype_semantics_xlsx.py"),
    ]
    for i, row in enumerate(readme, 1):
        for j, v in enumerate(row, 1):
            cell = ws.cell(i, j, v)
            cell.font = Font(name="Arial", size=14, bold=True, color="2F5496") if i == 1 else BODY_FONT
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100

    # --- Main semantics sheet ---
    ws2 = wb.create_sheet("AbilityType语义")
    headers = [
        "abilityType", "enumName", "代表法术中文名", "分类", "是否在SpellConfig",
        "float1语义", "float2语义", "float3语义", "int1语义", "int2语义", "int3语义",
        "配置样本值(Lv1-3)", "证据文件", "置信度", "备注",
    ]
    for c, h in enumerate(headers, 1):
        ws2.cell(1, c, h)

    # Ensure every enum id has a row
    all_ids = sorted(set(enum_map) | set(SEMANTICS))
    conf_count = defaultdict(int)
    for r_idx, ab in enumerate(all_ids, 2):
        enum_name = enum_map.get(ab, "?")
        if ab in SEMANTICS:
            cat, f1, f2, f3, i1, i2, i3, evid, conf, notes = SEMANTICS[ab]
        else:
            cat, f1, f2, f3, i1, i2, i3, evid, conf, notes = (
                "unknown", "unknown", "unknown", "unknown", "unknown", "unknown", "unknown",
                "SpellAbilityType.cs", "unknown", "未收录到手工语义表",
            )
        conf_count[conf] += 1
        row = [
            ab, enum_name, names.get(ab, ""), CATEGORY_CN.get(cat, cat),
            "YES" if ab in in_config else "NO",
            f1, f2, f3, i1, i2, i3,
            samples.get(ab, ""), evid, conf, notes,
        ]
        for c, v in enumerate(row, 1):
            cell = ws2.cell(r_idx, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r_idx % 2 == 0:
                cell.fill = ALT
            if c == 14 and conf in CONF_FILL:
                cell.fill = CONF_FILL[conf]
        ws2.row_dimensions[r_idx].height = 36

    style_header(ws2, len(headers))
    autosize(ws2)
    try:
        t = Table(displayName="T_AbilitySemantics", ref=f"A1:{get_column_letter(len(headers))}{ws2.max_row}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws2.add_table(t)
    except Exception:
        pass

    # --- Per-level samples ---
    ws3 = wb.create_sheet("按等级样本值")
    h3 = ["abilityType", "enumName", "spellId", "level", "name_zh",
          "float1", "float2", "float3", "int1", "int2", "int3",
          "damage", "mpCost", "speed", "duration", "radius", "useType", "dropType"]
    for c, h in enumerate(h3, 1):
        ws3.cell(1, c, h)
    tmap = {t["id"]: t for t in texts}
    r = 2
    for s in sorted(spells, key=lambda x: (x["abilityType"], x["id"])):
        ab = s["abilityType"]
        nm = tmap.get(s["id"] + 7_000_000, {}).get("chineseS", "") or ""
        vals = [
            ab, enum_map.get(ab, "?"), s["id"], s.get("level", 1), nm,
            s.get("float1", 0), s.get("float2", 0), s.get("float3", 0),
            s.get("int1", 0), s.get("int2", 0), s.get("int3", 0),
            s.get("damage", 0), s.get("mpCost", 0), s.get("speed", 0),
            s.get("duration", 0), s.get("radius", 0), s.get("useType", 0), s.get("dropType", 0),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(r, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = ALT
        r += 1
    style_header(ws3, len(h3))
    autosize(ws3)

    # --- Stats ---
    ws4 = wb.create_sheet("置信度统计")
    ws4["A1"] = "置信度"
    ws4["B1"] = "数量"
    ws4["C1"] = "说明"
    stats_desc = {
        "high": "代码或描述直接确认",
        "medium": "可判定未使用/间接证据",
        "low": "部分字段仍有猜测",
        "unknown": "证据不足或配置缺失",
    }
    for i, conf in enumerate(["high", "medium", "low", "unknown"], 2):
        ws4.cell(i, 1, conf).fill = CONF_FILL[conf]
        ws4.cell(i, 2, conf_count.get(conf, 0))
        ws4.cell(i, 3, stats_desc[conf])
        for c in range(1, 4):
            ws4.cell(i, c).font = BODY_FONT
            ws4.cell(i, c).border = THIN
    ws4.cell(6, 1, "枚举总数")
    ws4.cell(6, 2, len(all_ids))
    ws4.cell(7, 1, "SpellConfig中出现")
    ws4.cell(7, 2, len(in_config))
    ws4.cell(8, 1, "枚举有但配置无")
    ws4.cell(8, 2, len(set(all_ids) - in_config - {0}) if 0 in all_ids else len(set(all_ids) - in_config))
    for c in range(1, 3):
        ws4.cell(1, c).font = HEADER_FONT
        ws4.cell(1, c).fill = HEADER_FILL
    autosize(ws4)

    wb.save(OUT)
    print("Saved:", OUT)
    print("abilityTypes:", len(all_ids), "in SpellConfig:", len(in_config))
    print("confidence:", dict(conf_count))


if __name__ == "__main__":
    main()
