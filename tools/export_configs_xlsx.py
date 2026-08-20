# -*- coding: utf-8 -*-
"""Export all extracted Magicraft JSON configs into one multi-sheet .xlsx workbook."""
from __future__ import annotations

import json
import os
from copy import deepcopy
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = r"D:\SteamLibrary\steamapps\common\Magicraft\_reverse_engineering\assets_export"
OUT = r"D:\SteamLibrary\steamapps\common\Magicraft\_reverse_engineering\魔法工艺完整配置表.xlsx"

# (json filename without .json, sheet title <= 31 chars)
SHEETS = [
    ("SpellConfig", "SpellConfig"),
    ("WandConfig", "WandConfig"),
    ("RelicConfig", "RelicConfig"),
    ("CurseConfig", "CurseConfig"),
    ("SetConfig", "SetConfig"),
    ("RelicGroupConfig", "RelicGroupConfig"),
    ("PotionConfig", "PotionConfig"),
    ("ResearchConfig", "ResearchConfig"),
    ("SpecialObjConfig", "SpecialObjConfig"),
    ("UnitConfig", "UnitConfig"),
    ("TextConfig_Spell", "TextConfig_Spell"),
    ("TextConfig_Wand", "TextConfig_Wand"),
    ("TextConfig_Relic", "TextConfig_Relic"),
    ("TextConfig_RelicGroup", "TextConfig_RelicGroup"),
    ("TextConfig_Set", "TextConfig_Set"),
    ("TextConfig_Curse", "TextConfig_Curse"),
    ("TextConfig_Potion", "TextConfig_Potion"),
    ("TextConfig_Research", "TextConfig_Research"),
    ("TextConfig_Unit", "TextConfig_Unit"),
]

USE_TYPE = {0: "Missile", 1: "Summon", 2: "Enhance", 3: "Passive"}
DROP_TYPE = {0: "None", 1: "Common", 2: "Rare", 3: "Epic", 4: "Special"}

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
BODY_FONT = Font(name="Arial", size=10)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")


def load_json(name: str) -> list[dict]:
    path = os.path.join(BASE, f"{name}.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise TypeError(f"{name}: expected list, got {type(data)}")
    return data


def flatten_value(v: Any) -> Any:
    """Flatten nested dict/list into Excel-friendly scalars / JSON strings."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float, str)):
        return v
    if isinstance(v, dict):
        # Relic-style {value, valueUpgrade}
        if set(v.keys()) <= {"value", "valueUpgrade"} and "value" in v:
            return v
        return json.dumps(v, ensure_ascii=False, separators=(",", ":"))
    if isinstance(v, list):
        return json.dumps(v, ensure_ascii=False, separators=(",", ":"))
    return str(v)


def flatten_record(record: dict) -> dict:
    out: dict[str, Any] = {}
    for k, v in record.items():
        flat = flatten_value(v)
        if isinstance(flat, dict) and set(flat.keys()) <= {"value", "valueUpgrade"} and "value" in flat:
            out[f"{k}.value"] = flat.get("value", "")
            out[f"{k}.valueUpgrade"] = flat.get("valueUpgrade", "")
        else:
            out[k] = flat
    return out


def enrich_spell(rows: list[dict], texts: list[dict]) -> list[dict]:
    text_map = {t["id"]: t for t in texts}
    enriched = []
    for r in rows:
        e = deepcopy(r)
        tid_name = r["id"] + 7_000_000
        tid_desc = r["id"] + 7_100_000
        tname = text_map.get(tid_name, {})
        tdesc = text_map.get(tid_desc, {})
        e = {
            "id": r["id"],
            "name_zh": tname.get("chineseS", "") or "",
            "name_en": tname.get("english", "") or "",
            "useType": r.get("useType", ""),
            "useTypeName": USE_TYPE.get(r.get("useType"), ""),
            "abilityType": r.get("abilityType", ""),
            "level": r.get("level", ""),
            "dropType": r.get("dropType", ""),
            "dropTypeName": DROP_TYPE.get(r.get("dropType"), ""),
            "desc_zh": tdesc.get("chineseS", "") or "",
            "desc_en": tdesc.get("english", "") or "",
            **{k: v for k, v in r.items() if k not in ("id", "useType", "abilityType", "level", "dropType")},
        }
        enriched.append(flatten_record(e))
    return enriched


def enrich_wand(rows: list[dict], texts: list[dict]) -> list[dict]:
    text_map = {t["id"]: t for t in texts}
    enriched = []
    for r in rows:
        tid = r["id"] + 5_000_000
        t = text_map.get(tid, {})
        e = {
            "id": r["id"],
            "name_zh": t.get("chineseS", "") or "",
            "name_en": t.get("english", "") or "",
            **{k: v for k, v in r.items() if k != "id"},
        }
        enriched.append(flatten_record(e))
    return enriched


def enrich_named(rows: list[dict], texts: list[dict], id_offset: int) -> list[dict]:
    """Generic enrich with TextConfig where name id = id + offset."""
    text_map = {t["id"]: t for t in texts}
    enriched = []
    for r in rows:
        if "id" not in r:
            enriched.append(flatten_record(r))
            continue
        t = text_map.get(r["id"] + id_offset, {})
        e = {
            "id": r["id"],
            "name_zh": t.get("chineseS", "") or "",
            "name_en": t.get("english", "") or "",
            **{k: v for k, v in r.items() if k != "id"},
        }
        enriched.append(flatten_record(e))
    return enriched


def collect_headers(rows: list[dict]) -> list[str]:
    keys: list[str] = []
    seen = set()
    for r in rows:
        for k in r:
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:200]:  # sample first rows for speed
            if cell.value is None:
                continue
            length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(10, min(length + 2, max_width))


def write_sheet(wb: Workbook, title: str, rows: list[dict], *, replace_default: bool = False) -> None:
    if replace_default and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    if not rows:
        ws["A1"] = "(empty)"
        return

    headers = collect_headers(rows)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            val = row.get(h, "")
            if isinstance(val, bool):
                val = "TRUE" if val else "FALSE"
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 22
    autosize(ws)

    # Excel table (nice filters / style); skip if name collision risk on huge sheets
    table_name = "T_" + "".join(ch if ch.isalnum() else "_" for ch in title)[:20]
    try:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(headers))}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
    except Exception:
        pass


def write_readme(wb: Workbook, counts: dict[str, int]) -> None:
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "《魔法工艺》完整配置表"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="2F5496")
    lines = [
        "",
        "来源：Magicraft_Data/resources.assets → TextAsset (Configs/*)",
        "导出工具：_reverse_engineering/tools/export_configs_xlsx.py",
        "",
        "说明：",
        "1. 每个工作表对应一份 JSON 配置；嵌套对象已展平（如 Relic 的 int1.value / int1.valueUpgrade）。",
        "2. 数组字段（浮标槽位、锁定标记等）以 JSON 字符串保留，避免丢信息。",
        "3. SpellConfig / WandConfig / RelicConfig 等已附加 name_zh / name_en（来自对应 TextConfig）。",
        "4. SpellConfig 额外含 useTypeName、dropTypeName、desc_zh、desc_en。",
        "5. 文本表 TextConfig_* 保留全部语言列原文。",
        "",
        "各表行数：",
    ]
    for i, line in enumerate(lines, 2):
        ws.cell(i, 1, line).font = BODY_FONT
    start = 2 + len(lines)
    ws.cell(start, 1, "工作表").font = HEADER_FONT
    ws.cell(start, 1).fill = HEADER_FILL
    ws.cell(start, 2, "行数").font = HEADER_FONT
    ws.cell(start, 2).fill = HEADER_FILL
    for i, (name, n) in enumerate(counts.items(), start + 1):
        ws.cell(i, 1, name).font = BODY_FONT
        ws.cell(i, 2, n).font = BODY_FONT
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 12


def main() -> None:
    print("Loading JSON...")
    data = {name: load_json(name) for name, _ in SHEETS}

    spell_rows = enrich_spell(data["SpellConfig"], data["TextConfig_Spell"])
    wand_rows = enrich_wand(data["WandConfig"], data["TextConfig_Wand"])

    # Relic / Curse / Potion / Research / Set / RelicGroup / SpecialObj / Unit
    # Text id offsets from decompiled GetName patterns (best-effort):
    # Spell name = id+7000000, Wand name = id+5000000
    # Relic/others: try id itself first in TextConfig, then common offsets
    def try_enrich(rows, text_key, offsets):
        texts = data[text_key]
        text_ids = {t["id"] for t in texts}
        best = None
        best_hit = -1
        for off in offsets:
            hits = sum(1 for r in rows if "id" in r and (r["id"] + off) in text_ids)
            if hits > best_hit:
                best_hit = hits
                best = off
        if best is None or best_hit == 0:
            return [flatten_record(r) for r in rows]
        print(f"  {text_key}: using id offset +{best} (hits={best_hit}/{len(rows)})")
        return enrich_named(rows, texts, best)

    # Text id offsets verified against TextConfig_* (name = config.id + offset)
    COMMON_OFFSETS = [
        0, 1_000_000, 2_000_000, 3_000_000, 4_000_000, 5_000_000, 6_000_000,
        7_000_000, 8_000_000, 9_000_000, 10_000_000, 12_000_000, 15_000_000,
    ]
    relic_rows = try_enrich(data["RelicConfig"], "TextConfig_Relic", COMMON_OFFSETS)
    curse_rows = try_enrich(data["CurseConfig"], "TextConfig_Curse", COMMON_OFFSETS)
    potion_rows = try_enrich(data["PotionConfig"], "TextConfig_Potion", COMMON_OFFSETS)
    research_rows = try_enrich(data["ResearchConfig"], "TextConfig_Research", COMMON_OFFSETS)
    set_rows = try_enrich(data["SetConfig"], "TextConfig_Set", COMMON_OFFSETS)
    rgroup_rows = try_enrich(data["RelicGroupConfig"], "TextConfig_RelicGroup", COMMON_OFFSETS)
    special_rows = [flatten_record(r) for r in data["SpecialObjConfig"]]
    unit_rows = try_enrich(data["UnitConfig"], "TextConfig_Unit", COMMON_OFFSETS)

    text_sheets = {
        "TextConfig_Spell": [flatten_record(r) for r in data["TextConfig_Spell"]],
        "TextConfig_Wand": [flatten_record(r) for r in data["TextConfig_Wand"]],
        "TextConfig_Relic": [flatten_record(r) for r in data["TextConfig_Relic"]],
        "TextConfig_RelicGroup": [flatten_record(r) for r in data["TextConfig_RelicGroup"]],
        "TextConfig_Set": [flatten_record(r) for r in data["TextConfig_Set"]],
        "TextConfig_Curse": [flatten_record(r) for r in data["TextConfig_Curse"]],
        "TextConfig_Potion": [flatten_record(r) for r in data["TextConfig_Potion"]],
        "TextConfig_Research": [flatten_record(r) for r in data["TextConfig_Research"]],
        "TextConfig_Unit": [flatten_record(r) for r in data["TextConfig_Unit"]],
    }

    sheet_data = [
        ("SpellConfig", spell_rows),
        ("WandConfig", wand_rows),
        ("RelicConfig", relic_rows),
        ("CurseConfig", curse_rows),
        ("SetConfig", set_rows),
        ("RelicGroupConfig", rgroup_rows),
        ("PotionConfig", potion_rows),
        ("ResearchConfig", research_rows),
        ("SpecialObjConfig", special_rows),
        ("UnitConfig", unit_rows),
    ] + list(text_sheets.items())

    counts = {name: len(rows) for name, rows in sheet_data}

    print("Writing workbook...")
    wb = Workbook()
    # remove default after creating README
    default = wb.active
    wb.remove(default)

    write_readme(wb, counts)
    for name, rows in sheet_data:
        print(f"  sheet {name}: {len(rows)} rows")
        write_sheet(wb, name, rows)

    wb.save(OUT)
    print("Saved:", OUT)
    print("Total sheets:", len(wb.sheetnames))


if __name__ == "__main__":
    main()
