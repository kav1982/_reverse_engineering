# -*- coding: utf-8 -*-
"""Export SpellConfig.json (+ TextConfig_Spell / UnitConfig) into a full Excel workbook."""
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = r"D:\SteamLibrary\steamapps\common\Magicraft\_reverse_engineering\assets_export"
OUT = r"D:\SteamLibrary\steamapps\common\Magicraft\_reverse_engineering\魔法工艺_法术配置总表.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F3864")
SUB_FILL = PatternFill("solid", start_color="D9E1F2")
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USE_TYPE = {0: "主动弹幕 Missile", 1: "召唤 Summon", 2: "强化 Enhance", 3: "被动 Passive"}
DROP_TYPE = {0: "初始/不掉落 None", 1: "普通 Common", 2: "稀有 Rare", 3: "史诗 Epic", 4: "特殊 Special"}

ABILITY = {
    0: "None 无", 1000: "DefaultEmptySpell 空", 1001: "Bullet 魔法弹", 1002: "Rollball 滚球",
    1003: "Butterfly 蝴蝶", 1004: "Laser 激光", 1005: "PreFirework 烟花", 1006: "HoverTorch 悬浮火把",
    1007: "BlackHole 黑洞", 1008: "ArcaneExplosion 奥术爆炸", 1009: "BackMP 回蓝",
    1010: "SnakeWalk 蛇形弹", 1011: "DisintegrationRay 分解射线", 1012: "FireBall 火球",
    1013: "Meteor 陨石", 1014: "Rainbow 彩虹", 1015: "ArcaneNova 奥术新星", 1016: "Dash 冲刺",
    1017: "DeathAdder 死亡毒蛇", 1018: "ThunderAura 雷电光环", 1019: "HighPressureWasher 高压水枪",
    1020: "ManaCoin 魔力金币", 1021: "MagicBreaker 法术破坏者", 1022: "Boomerang 回旋镖",
    1023: "JudgementBlade 审判之刃", 1024: "GiantBubble 巨大气泡", 1025: "DragonBreath 龙息",
    1026: "ShiningStar 闪耀之星", 1027: "SuperNova 超新星", 1028: "MrBingArrow 冰先生之箭",
    1029: "DimensionTraveller 次元旅者", 1030: "Harpoons 鱼叉", 1031: "ShotGun 霰弹枪",
    2001: "Summon1 召唤物1", 2002: "Summon2 召唤物2(融合头)", 2003: "Summon3 召唤物3",
    2004: "Summon4 召唤物4(光柱)", 2005: "Summon5 召唤物5(法典书)", 2006: "Summon6 召唤物6",
    2007: "Summon7 召唤物7", 2008: "Summon8 召唤物8", 2009: "Summon9 召唤物9",
    3001: "Volley 齐射", 3002: "Multishot 多重射击", 3003: "TotalScattering 全方位散射",
    3004: "MucusCrystal 粘液水晶", 3005: "VenomCrystal 毒液水晶", 3006: "Penetrate 穿透",
    3007: "LightningChain 闪电链", 3008: "SpellHover 悬浮", 3009: "AroundOwner 环绕自身",
    3010: "AroundMouse 环绕鼠标", 3011: "FollowTarget 跟踪目标", 3012: "Rebound 反弹",
    3013: "SpellSplit 法术分裂", 3014: "Frozen 冰冻", 3015: "ParasiticWorm 寄生虫",
    3101: "ThunderCrystal 雷电水晶", 3102: "EnhanceAttackRatio 攻击力强化(加算)",
    3103: "EnhanceDurationValue 持续时间强化", 3104: "PowerSavingMode 节能模式(伤害乘算)",
    3105: "EnhanceCriticalChance 暴击率强化", 3106: "EnhanceSpeedValue 速度强化",
    3107: "EnhanceRadiusRatio 范围强化", 3108: "EnhanceSummonHPRecover 召唤物回血强化",
    3109: "Mimic 模拟(复制目标法术)", 3110: "LifeLine 生命线", 3111: "FireCrystal 火焰水晶(伤害加算)",
    3112: "PullForceCrystal 引力水晶", 3113: "RadiuRatioDown 范围衰减", 3114: "FollowOwner 跟随主人",
    3115: "FusionSummon 融合召唤物", 3116: "SpellEndTeleport 法术结束传送",
    3117: "RandomRotationRadiu 随机旋转半径", 3118: "TeammateSacrifice 队友献祭", 3119: "Fall 下落",
    3120: "TeammateSprite 队友精灵", 3121: "Refraction 折射(伤害乘算)", 3122: "Unyielding 不屈",
    3123: "OverDrive 超载(伤害加算)", 3124: "FatSpell 肥法术", 3125: "ReverseCast 反向施法",
    3126: "SpellLevelEnhance 法术等级强化", 3127: "SoulMate 灵魂伴侣",
    3128: "SpeedToDuration 速度转持续时间", 3129: "DeathInfect 死亡感染",
    3130: "RandomTeleport 随机传送", 3201: "OnOverTrigger 结束时触发器",
    3202: "OnOverSplitTrigger 结束分裂触发器", 3203: "OnMoveTrigger 移动触发器",
    3204: "OnStartRotationTrigger 起始旋转触发器", 3205: "OnHitTrigger 命中触发器",
    4001: "SacrificeBall 献祭球", 4002: "EmptyContainer 空的容器(+MP上限)",
    4003: "ManaEssence 魔力精华(+回蓝)", 4004: "ChargeMode 蓄力模式", 4005: "WandSpirit 法杖之魂",
    4006: "ForceCoolDown 强制冷却", 4007: "UltimateExtender 终极延伸器", 4008: "EchoRune 回声符文",
    4009: "ManaInterface 魔力接口", 4010: "AllFieldEnhance 全域强化",
    4011: "EqualDistributionAngle 等分角度", 4012: "Umbrella 保护伞", 4013: "RuneHammer 符文之锤",
    4014: "LaserBeam 激光束", 4015: "PostSlotExtenderMove 蓄能槽延伸器(移动)",
    4016: "PostSlotExtenderStand 蓄能槽延伸器(静止)", 4017: "PostSlotExtenderTime 蓄能槽延伸器(时间)",
    4018: "PostSlotExtenderCastSpell 蓄能槽延伸器(施法)", 4019: "BiAnLethalBlade 彼岸致命之刃",
    4020: "SpellEmbryo 法术胚胎", 4021: "ManaTendril 魔力卷须",
    4022: "RandomPosFocusMouse 随机位置聚焦鼠标", 4023: "ManaToPostChargeRatio 魔力转蓄能比",
    4024: "DaveHarpoons 戴夫鱼叉", 4025: "RedRune 红符文", 4026: "GreenRune 绿符文",
    4027: "BlueRune 蓝符文",
    9001: "BulletParabola 抛物线弹道", 9002: "BounceBone 弹跳骨头", 9003: "LongTrail 长拖尾",
    9004: "SoundWave 声波", 9005: "ChainStar 链星", 9006: "BulltHell 弹幕地狱",
    9007: "BulletSin 正弦弹道", 9008: "BulletSinSpeed 正弦速度弹道", 9009: "BladeWave 刀刃波",
    9010: "BounceBullet 弹跳子弹", 9011: "RotateArrow 旋转箭", 9012: "Bat 蝙蝠",
    9013: "BladeWaveVertical 竖直刀刃波", 9014: "Spear 长矛", 9015: "IcnBall 冰球",
    9016: "ChaseBullet 追踪子弹",
}

# (表头, JSON 字段, 类型, 含义)  —— 类型: i=整数 f=浮点 b=布尔 s=字符串 d=派生
COLUMNS = [
    ("ID", "id", "i", "法术唯一ID，常规法术满足 id = abilityType*10 + level"),
    ("中文名", "_name_cn", "d", "TextConfig_Spell 中 id+7000000 的 chineseS"),
    ("英文名", "_name_en", "d", "TextConfig_Spell 中 id+7000000 的 english"),
    ("家族ID", "_family", "d", "id//10，同一家族即同一法术的不同等级/变体，用于分组"),
    ("ID合规", "_id_rule_ok", "d", "id 是否等于 abilityType*10+level；False 多为敌方弹幕/换皮/占位条目"),
    ("等级", "level", "i", "法术等级，1/2/3；三个同名同级法术可三合一升级"),
    ("大类", "_use_type", "d", "useType 的可读名"),
    ("useType", "useType", "i", "SpellType 枚举：0弹幕 1召唤 2强化 3被动"),
    ("abilityType", "abilityType", "i", "SpellAbilityType 枚举，决定实际走哪套 ECS 逻辑/源码类"),
    ("能力类型名", "_ability", "d", "abilityType 的可读名"),
    ("稀有度", "_drop_type", "d", "dropType 的可读名"),
    ("dropType", "dropType", "i", "ItemDropType：0初始 1普通 2稀有 3史诗 4特殊"),
    ("金币价", "priceCoin", "i", "商店金币售价"),
    ("血量价", "priceHP", "i", "以生命值购买的价格（血商店）"),
    ("占用格数", "slotCost", "i", "在法杖槽位中占用的格子数"),
    ("槽位数修正", "slotNumModifyValue", "i", "装上后修改法杖可用槽位数（如 +1 格）"),
    ("蓝耗", "mpCost", "i", "单次施放消耗魔力"),
    ("伤害", "damage", "f", "基础伤害，进入 RatioValue 作为 BaseValue"),
    ("连发数", "shootCount", "i", "单次施放产生的弹体数量"),
    ("速度", "speed", "f", "弹体初速度"),
    ("持续", "duration", "f", "存在/持续时间（秒）"),
    ("半径", "radius", "f", "作用/碰撞半径"),
    ("暴击率", "criticalChance", "f", "暴击概率（百分数）"),
    ("击退", "knockback", "f", "命中击退力度"),
    ("后坐力", "recoil", "f", "对施法者自身的后坐力"),
    ("散射角", "angle", "f", "散射/偏转角度（度）"),
    ("上抛速度", "upSpeed", "f", "Y 轴初速度，配合 gravity 形成抛物线"),
    ("重力", "gravity", "f", "重力加速度"),
    ("射击间隔修正", "shootIntervalAddSubRevise", "f", "法杖射击间隔加减修正（秒）"),
    ("冷却加减", "coolDownAddSubRevise", "f", "法杖冷却加减修正（秒）"),
    ("冷却倍率", "coolDownRatio", "f", "法杖冷却乘算倍率，1=不变"),
    ("蓝耗加减修正", "mpCostAddSubCorrection", "f", "对后续法术蓝耗的加减修正"),
    ("蓝耗乘除修正", "mpCostMulDivCorrection", "f", "对后续法术蓝耗的乘除修正（百分数）"),
    ("isDPS", "isDPS", "b", "是否为持续伤害型（按间隔结算）"),
    ("DPS间隔", "DPSDamageInterval", "f", "持续伤害的结算间隔（秒）"),
    ("持续施法", "isKeepCasting", "b", "是否需要按住施法键引导"),
    ("可取消引导", "canCancelCasting", "b", "引导过程中能否中断"),
    ("父法术", "isParentTypeSpell", "b", "是否为容器型父法术（可携带子法术）"),
    ("对弹幕生效", "haveEffecforMissileSpell", "b", "强化效果是否作用于 Missile 类法术"),
    ("对召唤生效", "haveEffecforSummonSpell", "b", "强化效果是否作用于 Summon 类法术"),
    ("分裂法术", "isSplitSpell", "b", "是否由分裂机制产生的子法术"),
    ("可合成", "canCompound", "b", "是否参与三合一合成"),
    ("需激活", "needActivate", "b", "是否需要满足条件后才生效"),
    ("summonID", "summonID", "i", "召唤的单位 ID，对应 UnitConfig"),
    ("召唤物名", "_summon_name", "d", "UnitConfig 对应单位名（id+12000000）"),
    ("召唤上限", "summonLimit", "i", "同时存在的召唤物上限"),
    ("召唤跟随过关", "summonCanEnterNextMap", "b", "召唤物能否跟随进入下一层"),
    ("float1", "float1", "f", "通用浮点参数，语义随 abilityType 变化（见描述模板占位符）"),
    ("float2", "float2", "f", "同上"),
    ("float3", "float3", "f", "同上"),
    ("int1", "int1", "i", "通用整型参数，语义随 abilityType 变化（见描述模板占位符）"),
    ("int2", "int2", "i", "同上"),
    ("int3", "int3", "i", "同上"),
    ("效果描述模板", "_desc_tpl", "d", "id+7100000 原文，内含 float1/int1 等占位符，是解读通用字段的关键"),
    ("效果描述(代入数值)", "_desc_val", "d", "把模板占位符替换为本行实际数值（近似，不含运行时加成）"),
    ("辅助说明", "_assist", "d", "(id//10*10+1)+7200000"),
    ("简介", "_intro", "d", "(id//10*10+1)+7300000"),
    ("英文描述模板", "_desc_en", "d", "id+7100000 的 english"),
    ("图标", "icon", "s", "图标资源名"),
    ("iconH", "iconH", "s", "高亮/特殊状态图标名"),
    ("prefab", "prefab", "s", "运行时预制体资源名"),
    ("播放音效", "playShootSE", "b", "施放时是否播放音效"),
]

PLACEHOLDERS = ["float1", "float2", "float3", "int1", "int2", "int3"]


def load(name):
    with open(os.path.join(BASE, name + ".json"), encoding="utf-8") as fp:
        return json.load(fp)


def fmt_num(v):
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def main():
    spells = sorted(load("SpellConfig"), key=lambda s: (s["useType"], s["id"] // 10, s["level"], s["id"]))
    texts = {t["id"]: t for t in load("TextConfig_Spell")}
    unit_texts = {t["id"]: t for t in load("TextConfig_Unit")}

    def txt(tid, lang="chineseS"):
        return (texts.get(tid) or {}).get(lang) or ""

    def enrich(s):
        base = s["id"] // 10 * 10 + 1
        tpl = txt(s["id"] + 7100000)
        val = tpl
        for ph in PLACEHOLDERS:
            if ph in val:
                val = val.replace(ph, fmt_num(s.get(ph, 0)))
        s["_family"] = s["id"] // 10
        s["_id_rule_ok"] = s["id"] == s["abilityType"] * 10 + s["level"]
        s["_name_cn"] = txt(s["id"] + 7000000)
        s["_name_en"] = txt(s["id"] + 7000000, "english")
        s["_use_type"] = USE_TYPE.get(s["useType"], str(s["useType"]))
        s["_ability"] = ABILITY.get(s["abilityType"], str(s["abilityType"]))
        s["_drop_type"] = DROP_TYPE.get(s["dropType"], str(s["dropType"]))
        s["_desc_tpl"] = tpl
        s["_desc_val"] = val
        s["_assist"] = txt(base + 7200000).replace("\\", "\n")
        s["_intro"] = txt(base + 7300000)
        s["_desc_en"] = txt(s["id"] + 7100000, "english")
        s["_summon_name"] = (unit_texts.get(s["summonID"] + 12000000) or {}).get("chineseS") or "" if s["summonID"] else ""
        return s

    spells = [enrich(s) for s in spells]

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True  # 打开时强制重算跨表公式

    # ---------------- 法术总表 ----------------
    ws = wb.active
    ws.title = "法术总表"
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for s in spells:
        ws.append([s.get(c[1]) for c in COLUMNS])
    style_header(ws, len(headers))

    type_fill = {
        0: PatternFill("solid", start_color="FFF2CC"),
        1: PatternFill("solid", start_color="E2EFDA"),
        2: PatternFill("solid", start_color="DEEBF7"),
        3: PatternFill("solid", start_color="FCE4D6"),
    }
    use_col = headers.index("大类") + 1
    text_cols = {headers.index(h) + 1 for h in
                 ("效果描述模板", "效果描述(代入数值)", "辅助说明", "简介", "英文描述模板")}
    for r, s in enumerate(spells, start=2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in text_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=use_col).fill = type_fill[s["useType"]]

    widths = {"ID": 8, "中文名": 16, "英文名": 22, "等级": 5, "大类": 17, "useType": 8,
              "abilityType": 11, "能力类型名": 30, "稀有度": 15, "dropType": 9,
              "效果描述模板": 42, "效果描述(代入数值)": 42, "辅助说明": 40, "简介": 32,
              "英文描述模板": 42, "召唤物名": 14, "prefab": 12, "图标": 10, "iconH": 10}
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, max(7, min(14, len(h) * 2 + 3)))
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(spells) + 1}"

    MAIN = "'法术总表'"
    col_of = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

    # ---------------- 等级成长对比 ----------------
    ws2 = wb.create_sheet("等级成长对比")
    h2 = ["家族ID", "中文名", "大类", "abilityType", "能力类型名", "稀有度", "记录数",
          "Lv1 ID", "Lv2 ID", "Lv3 ID",
          "Lv1伤害", "Lv2伤害", "Lv3伤害", "伤害成长(Lv3/Lv1)",
          "Lv1蓝耗", "Lv2蓝耗", "Lv3蓝耗",
          "Lv1金币", "Lv2金币", "Lv3金币", "备注"]
    ws2.append(h2)
    ID_COL = {1: 8, 2: 9, 3: 10}
    METRIC_COL = ((11, "伤害"), (15, "蓝耗"), (18, "金币价"))
    families = {}
    for s in spells:
        families.setdefault(s["_family"], []).append(s)

    rows = []
    for family, group in sorted(families.items(), key=lambda kv: (kv[1][0]["useType"], kv[0])):
        group = sorted(group, key=lambda x: (x["level"], x["id"]))
        slots, extras = {}, []
        for g in group:
            if g["level"] in ID_COL and g["level"] not in slots:
                slots[g["level"]] = g
            else:
                extras.append(g)
        head = group[0]
        abilities = sorted({g["abilityType"] for g in group})
        note = "" if all(g["_id_rule_ok"] for g in group) else "ID 不符合 abilityType*10+level（敌方弹幕/换皮/占位条目）"
        rows.append((family, head, slots, abilities, len(group), note))
        for ex in extras:
            rows.append((family, ex, {ex["level"]: ex} if ex["level"] in ID_COL else {}, [ex["abilityType"]], 1,
                         "同家族额外变体，另计一行"))

    r = 2
    for family, head, slots, abilities, count, note in rows:
        ws2.cell(row=r, column=1, value=family)
        ws2.cell(row=r, column=2, value=head["_name_cn"])
        ws2.cell(row=r, column=3, value=head["_use_type"])
        ws2.cell(row=r, column=4, value="/".join(str(a) for a in abilities))
        ws2.cell(row=r, column=5, value=" / ".join(ABILITY.get(a, str(a)) for a in abilities))
        ws2.cell(row=r, column=6, value=head["_drop_type"])
        ws2.cell(row=r, column=7, value=count)
        for lv, c in ID_COL.items():
            if lv in slots:
                ws2.cell(row=r, column=c, value=slots[lv]["id"])
        for offset, metric in METRIC_COL:
            src = col_of[metric]
            for j, lv in enumerate((1, 2, 3)):
                key_cell = f"{get_column_letter(ID_COL[lv])}{r}"
                ws2.cell(row=r, column=offset + j).value = (
                    f'=IF({key_cell}="","",IFERROR(INDEX({MAIN}!${src}:${src},'
                    f'MATCH({key_cell},{MAIN}!$A:$A,0)),""))')
        ws2.cell(row=r, column=14).value = f'=IF(OR($K{r}="",$K{r}=0,$M{r}=""),"",$M{r}/$K{r})'
        ws2.cell(row=r, column=21, value=note)
        r += 1
    style_header(ws2, len(h2))
    for rr in range(2, r):
        for c in range(1, len(h2) + 1):
            cell = ws2.cell(row=rr, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row=rr, column=14).number_format = "0.00x"
        ws2.cell(row=rr, column=21).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, h in enumerate(h2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = {"能力类型名": 32, "大类": 17, "中文名": 16,
                                                             "稀有度": 15, "伤害成长(Lv3/Lv1)": 15,
                                                             "abilityType": 12, "备注": 46}.get(h, 9)
    ws2.freeze_panes = "D2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{r - 1}"

    # ---------------- 字段字典 ----------------
    ws3 = wb.create_sheet("字段字典")
    ws3.append(["总表列", "列头", "JSON 字段", "数据类型", "含义 / 来源说明"])
    type_name = {"i": "int", "f": "float", "b": "bool", "s": "string", "d": "派生(非原始字段)"}
    for i, (head, field, typ, desc) in enumerate(COLUMNS):
        ws3.append([get_column_letter(i + 1), head, "" if typ == "d" else field, type_name[typ], desc])
    style_header(ws3, 5)
    for rr in range(2, len(COLUMNS) + 2):
        for c in range(1, 6):
            cell = ws3.cell(row=rr, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 5))
        if ws3.cell(row=rr, column=4).value == "派生(非原始字段)":
            for c in range(1, 6):
                ws3.cell(row=rr, column=c).fill = SUB_FILL
    for c, w in zip("ABCDE", (8, 22, 26, 16, 70)):
        ws3.column_dimensions[c].width = w
    ws3.freeze_panes = "A2"

    # ---------------- 枚举对照 ----------------
    ws4 = wb.create_sheet("枚举对照")
    ws4["A1"] = "SpellConfig 相关枚举对照（来自 Assembly-CSharp 反编译）"
    ws4["A1"].font = TITLE_FONT
    row = 3
    header_rows = []
    ws4.cell(row=row, column=1, value="SpellType (useType)")
    ws4.cell(row=row, column=2, value="可读名")
    ws4.cell(row=row, column=3, value="总表条数")
    style_header(ws4, 3, row)
    header_rows.append(row)
    row += 1
    for k, v in USE_TYPE.items():
        ws4.cell(row=row, column=1, value=k)
        ws4.cell(row=row, column=2, value=v)
        ws4.cell(row=row, column=3,
                 value=f'=COUNTIF({MAIN}!${col_of["useType"]}:${col_of["useType"]},A{row})')
        row += 1
    row += 1
    ws4.cell(row=row, column=1, value="ItemDropType (dropType)")
    ws4.cell(row=row, column=2, value="可读名")
    ws4.cell(row=row, column=3, value="总表条数")
    style_header(ws4, 3, row)
    header_rows.append(row)
    row += 1
    for k, v in DROP_TYPE.items():
        ws4.cell(row=row, column=1, value=k)
        ws4.cell(row=row, column=2, value=v)
        ws4.cell(row=row, column=3,
                 value=f'=COUNTIF({MAIN}!${col_of["dropType"]}:${col_of["dropType"]},A{row})')
        row += 1
    row += 1
    ws4.cell(row=row, column=1, value="SpellAbilityType")
    ws4.cell(row=row, column=2, value="可读名")
    ws4.cell(row=row, column=3, value="总表条数")
    style_header(ws4, 3, row)
    header_rows.append(row)
    row += 1
    for k, v in sorted(ABILITY.items()):
        ws4.cell(row=row, column=1, value=k)
        ws4.cell(row=row, column=2, value=v)
        ws4.cell(row=row, column=3,
                 value=f'=COUNTIF({MAIN}!${col_of["abilityType"]}:${col_of["abilityType"]},A{row})')
        row += 1
    for rr in range(3, row):
        for c in range(1, 4):
            cell = ws4.cell(row=rr, column=c)
            if rr not in header_rows:
                cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="center" if c != 2 else "left", vertical="center")
            cell.border = BORDER
    for c, w in zip("ABC", (24, 40, 12)):
        ws4.column_dimensions[c].width = w

    # ---------------- 说明 ----------------
    ws0 = wb.create_sheet("说明", 0)
    ws0["A1"] = "《魔法工艺 Magicraft》法术配置总表"
    ws0["A1"].font = TITLE_FONT
    info = [
        ("数据来源", "Magicraft_Data/resources.assets 内 TextAsset：Resources/Configs/SpellConfig.json"),
        ("文本来源", "Resources/Configs/TextConfig_Spell.json、TextConfig_Unit.json"),
        ("提取方式", "UnityPy 导出 TextAsset → JSON；字段语义由 ilspycmd 反编译 Assembly-CSharp.dll 交叉验证"),
        ("记录条数", '=COUNT(\'法术总表\'!$A$2:$A$100000) & " 条（含同一法术的 Lv1/Lv2/Lv3 独立记录）"'),
        ("法术家族数", '=COUNTA(\'等级成长对比\'!$A$2:$A$10000) & " 个家族（按 id//10 归并等级链）"'),
        ("ID 规则", "常规法术 id = abilityType*10 + level；同一法术的 Lv1/2/3 是三条独立记录"),
        ("ID 例外", f'=COUNTIF({MAIN}!${col_of["ID合规"]}:${col_of["ID合规"]},FALSE) & " 条不满足该规则（敌方弹幕、换皮变体与 20000/30000/40000/90000 占位空法术），见总表“ID合规”列"'),
        ("数值口径", "damage 等为配置基础值，实战值走 RatioValue：(BaseValue+addBase)*baseAddRatio*addRatio*mulRatio+addExtra"),
        ("通用字段", "float1~3 / int1~3 语义随 abilityType 变化，请对照“效果描述模板”列中的占位符解读"),
        ("", ""),
        ("工作表", "内容"),
        ("法术总表", "全部法术的原始字段值 + 名称/描述/枚举可读名（已开启筛选，冻结前三列）"),
        ("等级成长对比", "按法术家族汇总 Lv1/2/3 的 ID、伤害、蓝耗、金币价与伤害成长倍率（公式引用总表，改总表即自动更新）"),
        ("字段字典", "总表每一列对应的 JSON 字段、数据类型与含义"),
        ("枚举对照", "SpellType / ItemDropType / SpellAbilityType 枚举值与中文名，附各值条数统计"),
    ]
    r0 = 3
    for k, v in info:
        ws0.cell(row=r0, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        ws0.cell(row=r0, column=2, value=v).font = BODY_FONT
        ws0.cell(row=r0, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        r0 += 1
    for rr, (k, v) in enumerate(info, start=3):
        if k == "工作表":
            for c in (1, 2):
                ws0.cell(row=rr, column=c).fill = SUB_FILL
    ws0.column_dimensions["A"].width = 16
    ws0.column_dimensions["B"].width = 110

    wb.save(OUT)
    print("saved", OUT, "spells", len(spells), "families", len(families), "rows", len(rows))


if __name__ == "__main__":
    main()
